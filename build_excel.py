import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import csv

wb = openpyxl.Workbook()

# ─── Colour palette ───────────────────────────────────────────────────────────
DARK_BLUE   = "1A2E4A"
MID_BLUE    = "2E5FA3"
LIGHT_BLUE  = "D6E4F7"
ORANGE      = "E8600A"
LIGHT_ORANGE= "FDEBD0"
GREEN       = "1E7E34"
LIGHT_GREEN = "D4EDDA"
RED         = "C0392B"
LIGHT_RED   = "FADBD8"
YELLOW_BG   = "FFF9C4"
GREY_HDR    = "F2F2F2"
WHITE       = "FFFFFF"

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def body_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — EDITORIAL CALENDAR
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Q2 Editorial Calendar"

# --- read CSV -----------------------------------------------------------------
with open("/home/user/my-web-app/q2_editorial_calendar.csv", newline="", encoding="utf-8") as f:
    rows = list(csv.reader(f))

headers = rows[0]
data    = [r for r in rows[1:] if any(r)]

# --- header row ---------------------------------------------------------------
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font      = hdr_font()
    cell.fill      = fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border()

ws1.row_dimensions[1].height = 30

# --- data rows ----------------------------------------------------------------
URGENT_FILL   = fill(LIGHT_RED)
HIGH_FILL     = fill(LIGHT_ORANGE)
PILLAR_FILL   = fill(LIGHT_BLUE)

priority_col  = headers.index("Priority")   # 18 (0-based)
class_col     = headers.index("Class")      # 8

for row_idx, row in enumerate(data, 2):
    priority = row[priority_col] if len(row) > priority_col else ""
    cls      = row[class_col]    if len(row) > class_col    else ""

    for col_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = body_font()
        cell.border    = border()
        cell.alignment = wrap_align()

        # row tinting
        if priority == "URGENT":
            cell.fill = URGENT_FILL
        elif cls == "Pillar":
            cell.fill = PILLAR_FILL
        elif priority == "High":
            cell.fill = fill(GREY_HDR)
        else:
            cell.fill = fill(WHITE)

    ws1.row_dimensions[row_idx].height = 40

# --- column widths ------------------------------------------------------------
col_widths = {
    "Campaign": 12, "Assignee": 16, "Delivery deadline": 14,
    "Proofreader": 16, "Proofread deadline": 14, "Status": 14,
    "Publication date": 14, "Topic": 42, "Class": 10, "Funnel": 8,
    "Service-line": 20, "SEO Keyword": 28, "Angle": 44,
    "Market": 8, "Location": 14, "Language": 18,
    "Category name": 16, "Priority": 10, "Note": 48,
}
for col_idx, h in enumerate(headers, 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 14)

ws1.freeze_panes = "A2"

# --- legend -------------------------------------------------------------------
legend_row = len(data) + 4
ws1.cell(row=legend_row,   column=1, value="LEGEND").font = hdr_font(color=DARK_BLUE)
items = [
    (LIGHT_RED,    "URGENT — publish immediately"),
    (LIGHT_BLUE,   "Pillar article"),
    (GREY_HDR,     "High priority cluster"),
    (WHITE,        "Medium priority cluster"),
]
for i, (clr, label) in enumerate(items):
    r = legend_row + 1 + i
    ws1.cell(row=r, column=1).fill  = fill(clr)
    ws1.cell(row=r, column=1).border = border()
    ws1.cell(row=r, column=2, value=label).font = body_font()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — STRATEGY RATIONALE
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Q2 Strategy Rationale")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 90

def section_hdr(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font  = Font(bold=True, color=WHITE, size=12, name="Calibri")
    cell.fill  = fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 22
    return row + 1

def label_value(ws, row, label, value, label_bold=True, val_fill=None):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font      = Font(bold=label_bold, size=10, name="Calibri", color=DARK_BLUE)
    lc.fill      = fill(LIGHT_BLUE)
    lc.alignment = wrap_align("left", "top")
    lc.border    = border()

    vc = ws.cell(row=row, column=2, value=value)
    vc.font      = body_font()
    vc.alignment = wrap_align()
    vc.border    = border()
    if val_fill:
        vc.fill = fill(val_fill)

    ws.row_dimensions[row].height = max(15, min(120, len(str(value)) // 2))
    return row + 1

r = 1

# Title
title = ws2.cell(row=r, column=1, value="GuestReady — Q2 2026 Content Strategy: Rationale")
title.font = Font(bold=True, size=16, color=DARK_BLUE, name="Calibri")
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.row_dimensions[r].height = 30
r += 2

# ── 1. Context ────────────────────────────────────────────────────────────────
r = section_hdr(ws2, r, "1. Strategic Context")
r = label_value(ws2, r, "Where we are",
    "Q1 established foundational content across all 5 markets (UK, FR, ES, PT, UAE). "
    "City-level cluster articles are live for London and Manchester (UK), Paris and Lyon (FR), "
    "Málaga and Valencia (ES), Lisbon and Porto (PT), and Dubai (UAE). "
    "Several Q1 items carried forward due to resourcing and staging delays.")
r = label_value(ws2, r, "What Q2 must achieve",
    "1. Complete the cluster series so pillar articles can publish with full internal linking. "
    "2. Capitalise on two urgent regulatory moments (Paris Airbnb fin, Andalucía licence). "
    "3. Open the mid-term rental angle in Paris and UAE. "
    "4. Begin building conversion-focused pillar pages that link clusters and drive leads.")
r += 1

# ── 2. Market decisions ───────────────────────────────────────────────────────
r = section_hdr(ws2, r, "2. Market-by-Market Decisions")

r = label_value(ws2, r, "🇫🇷 France (7 articles)",
    "Heaviest investment this quarter. Paris regulatory landscape is shifting fast — "
    "'fin Airbnb Paris 2026' is an urgent TOFU capture piece. "
    "The mid-term rental series (bail mobilité vs saisonnière → passer LCD à MLD) builds a "
    "thematic cluster that directly supports GuestReady's mid-term product pivot. "
    "The FR conciergerie pillar (May 12) caps the Lyon + Paris cluster work from Q1 and Q2.")

r = label_value(ws2, r, "🇬🇧 UK (6 articles)",
    "Edinburgh, Brighton, and Liverpool city PM cluster articles complete the series started with "
    "London + Manchester in Q1. All three must be live before the UK Pillar (Jun 2) to enable "
    "full internal linking. Dublin content (PM companies + renting out property) addresses a "
    "high-value adjacent market. Holiday let fee comparison article builds E-E-A-T trust signals.")

r = label_value(ws2, r, "🇪🇸 Spain (4 articles)",
    "Andalucía pillar is already staged — just needs publishing (URGENT). "
    "Canarias regulatory pillar (May 22) extends the ES regulatory series. "
    "The ES PM companies pillar (Jun 9) ties together Andalucía + Canarias + Q1 Valencia/Málaga clusters. "
    "The expat property management cluster targets the ES02/ES04/ES05 remote-owner persona with no existing content.")

r = label_value(ws2, r, "🇵🇹 Portugal (4 articles)",
    "PT receives lighter investment this quarter given Q1 carry-forwards. "
    "AL investment cluster and alternativa a AL pillar address the regulatory anxiety angle "
    "driven by Portuguese AL law changes. Madeira cluster (Jun 30) extends geographic coverage. "
    "Emigrant remote management cluster is a high-differentiation angle with no competitor content.")

r = label_value(ws2, r, "🇦🇪 UAE (3 articles)",
    "Mid-term Dubai cluster mirrors the successful FR mid-term pattern. "
    "The Dubai investor pillar (Jun 23) umbrellas the nationality-based investment articles from Q1 "
    "(Saudi, Pakistani, UK). Staycation cluster addresses a dominant 2025–26 Dubai search trend.")
r += 1

# ── 3. Content architecture ───────────────────────────────────────────────────
r = section_hdr(ws2, r, "3. Content Architecture")
r = label_value(ws2, r, "Pillar-Cluster model",
    "Every pillar article publishes AFTER its supporting cluster articles are live. "
    "This ensures full internal linking at launch, which is the primary ranking mechanism for pillar pages. "
    "Sequencing is enforced in the calendar via the 'Note' column dependency flags.")
r = label_value(ws2, r, "Funnel balance",
    "TOFU: 1 article (4%) — urgency-driven regulatory capture.\n"
    "MOFU: 17 articles (65%) — owner education, comparison, and decision-stage content.\n"
    "BOFU: 8 articles (31%) — high-intent 'best PM companies' and conciergerie queries.")
r = label_value(ws2, r, "Class split",
    "Pillar: 7 articles — high DA/PA targets, extensive internal linking, 2,500+ words.\n"
    "Cluster: 19 articles — supporting evidence for pillar pages, 1,000–1,800 words.")
r += 1

# ── 4. Carry-forwards ─────────────────────────────────────────────────────────
r = section_hdr(ws2, r, "4. Q1 Carry-Forwards Included in Q2")
carries = [
    ("Edinburgh PM companies", "Apr 10", "UK Pillar dependency"),
    ("Dublin PM companies", "Apr 17", "Pairs with Dublin regulations May"),
    ("Lyon conciergerie Airbnb", "Apr 22", "FR Pillar dependency"),
    ("Logement flexible Paris", "Apr 24", "FR mid-term series"),
    ("Investir em alojamento local PT", "Apr 28", "PT regulatory angle"),
    ("Brighton PM companies", "May 8", "UK Pillar dependency"),
    ("Liverpool PM companies", "May 15", "UK Pillar dependency"),
    ("Alternativa a alojamento local PT", "Jun 12", "PT pillar"),
    ("Louer au mois Airbnb FR", "Jun 19", "FR mid-term series"),
    ("Empresas AL Madeira", "Jun 30", "PT geographic expansion"),
]
for item, date, reason in carries:
    r = label_value(ws2, r, item, f"Rescheduled to {date}. Reason: {reason}")
r += 1

# ── 5. What's deferred to Q3 ─────────────────────────────────────────────────
r = section_hdr(ws2, r, "5. Deferred to Q3 (not in Q2 calendar)")
r = label_value(ws2, r, "Paris regulatory mega-pillar",
    "Requires full Airbnb fin article (Q2 cluster) + bail mobilité + bail code civil to be live first. "
    "Estimated Q3 delivery: July 2026.")
r = label_value(ws2, r, "Case studies (all markets)",
    "Dependent on client content sign-off. Not suitable for editorial calendar until source material confirmed.")
r = label_value(ws2, r, "Q1 ideas not yet briefed",
    "Several Q2 ideas from the original Q2 list were deprioritised due to: no existing cluster support, "
    "low urgency, or overlap with already-live content.")

ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — KPIs
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Q2 KPIs")
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 50

r = 1
title3 = ws3.cell(row=r, column=1, value="GuestReady — Q2 2026 Content KPIs")
title3.font = Font(bold=True, size=16, color=DARK_BLUE, name="Calibri")
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws3.row_dimensions[r].height = 30
r += 1

sub = ws3.cell(row=r, column=1, value="Measurement period: 1 April – 30 June 2026 | Review date: 15 July 2026")
sub.font = Font(italic=True, size=10, color="666666", name="Calibri")
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 2

# ── Section header helper ─────────────────────────────────────────────────────
def kpi_section(ws, row, text):
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(DARK_BLUE)
        cell.border = border()
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws.row_dimensions[row].height = 20
    return row + 1

def kpi_col_hdr(ws, row):
    hdrs = ["KPI", "Baseline (end Q1)", "Q2 Target", "Stretch Target", "How to measure"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font      = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
        cell.fill      = fill(LIGHT_BLUE)
        cell.border    = border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    return row + 1

def kpi_row(ws, row, kpi, baseline, target, stretch, method, tint=None):
    vals = [kpi, baseline, target, stretch, method]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font      = body_font(size=10)
        cell.border    = border()
        cell.alignment = wrap_align("left" if ci in (1,5) else "center")
        if tint:
            cell.fill = fill(tint)
    ws.row_dimensions[row].height = max(18, min(60, len(str(method)) // 3))
    return row + 1

# ── 1. Output KPIs ────────────────────────────────────────────────────────────
r = kpi_section(ws3, r, "1. OUTPUT — What we publish")
r = kpi_col_hdr(ws3, r)
r = kpi_row(ws3, r, "Articles published on schedule",     "—",     "24 / 26",  "26 / 26",  "Editorial calendar status column", LIGHT_GREEN)
r = kpi_row(ws3, r, "Pillar articles live",               "0 Q2",  "7",        "7",        "WordPress publish log")
r = kpi_row(ws3, r, "Cluster articles live",              "—",     "19",       "19",       "WordPress publish log")
r = kpi_row(ws3, r, "Markets with ≥1 new pillar",         "0",     "5 / 5",    "5 / 5",    "One pillar per market: UK/FR/ES/PT/UAE")
r = kpi_row(ws3, r, "On-time delivery rate",              "—",     "85%",      "95%",      "Delivery deadline vs actual publish date")
r += 1

# ── 2. SEO performance KPIs ───────────────────────────────────────────────────
r = kpi_section(ws3, r, "2. SEO PERFORMANCE — Organic visibility")
r = kpi_col_hdr(ws3, r)
r = kpi_row(ws3, r, "Total organic sessions (new Q2 articles)", "0",      "+4,000 sessions",  "+7,000 sessions",  "GA4: organic channel, new pages only — 90-day window", LIGHT_GREEN)
r = kpi_row(ws3, r, "Pages ranking top 10 (target keyword)",    "Measure at launch", "8 / 26 pages",  "14 / 26 pages", "GSC: position filter ≤10 on primary keyword")
r = kpi_row(ws3, r, "Pages ranking top 20 (target keyword)",    "—",      "18 / 26 pages",    "22 / 26 pages",    "GSC: position filter ≤20 on primary keyword")
r = kpi_row(ws3, r, "Average position — pillar articles",       "—",      "≤18",              "≤12",              "GSC: average position filter on 7 pillar URLs", LIGHT_BLUE)
r = kpi_row(ws3, r, "Indexed within 14 days of publish",        "—",      "90% of new pages", "100%",             "GSC: Coverage report → Valid pages")
r = kpi_row(ws3, r, "Internal links pointing to each pillar",   "0",      "≥4 per pillar",    "≥6 per pillar",    "Screaming Frog crawl or Ahrefs Site Audit")
r += 1

# ── 3. AEO / AI visibility ────────────────────────────────────────────────────
r = kpi_section(ws3, r, "3. AEO — AI answer engine visibility")
r = kpi_col_hdr(ws3, r)
r = kpi_row(ws3, r, "Pages with FAQ schema implemented",         "—",   "26 / 26",   "26 / 26",   "Technical audit: JSON-LD FAQPage present", LIGHT_ORANGE)
r = kpi_row(ws3, r, "Pages with Speakable schema implemented",   "—",   "26 / 26",   "26 / 26",   "Technical audit: JSON-LD Speakable present")
r = kpi_row(ws3, r, "GSC 'rich results' FAQ appearances",        "Baseline Q1", "+15%",     "+30%",      "GSC Enhancements → FAQ results")
r = kpi_row(ws3, r, "Cited in Perplexity / ChatGPT answers",    "0 tracked", "Track & log", "2 citations", "Manual spot-check on target queries monthly", LIGHT_ORANGE)
r += 1

# ── 4. Conversion KPIs ────────────────────────────────────────────────────────
r = kpi_section(ws3, r, "4. CONVERSION — Business impact")
r = kpi_col_hdr(ws3, r)
r = kpi_row(ws3, r, "Owner enquiries from organic (all markets)",   "Establish Q1 baseline", "+10% QoQ",  "+20% QoQ",  "GA4: goal 'Owner enquiry form' × organic source", LIGHT_GREEN)
r = kpi_row(ws3, r, "Click-throughs to owner landing pages",        "—",     "+15% QoQ",  "+25% QoQ",  "GA4: event 'owner_page_click' from blog posts")
r = kpi_row(ws3, r, "Time on page — pillar articles",               "—",     "≥2:30 avg", "≥3:30 avg", "GA4: engagement time, pillar URL filter")
r = kpi_row(ws3, r, "Bounce rate — new Q2 articles",                "—",     "≤70%",      "≤60%",      "GA4: bounce rate, new pages segment")
r += 1

# ── 5. Quality KPIs ───────────────────────────────────────────────────────────
r = kpi_section(ws3, r, "5. QUALITY — Editorial standards")
r = kpi_col_hdr(ws3, r)
r = kpi_row(ws3, r, "Articles passing proofreader review first pass", "—", "80%",  "90%",  "Proofreader marks 'approved' without revision round", LIGHT_BLUE)
r = kpi_row(ws3, r, "Articles with structured data errors at launch",  "—", "0",   "0",    "Google Rich Results Test at publish")
r = kpi_row(ws3, r, "Articles missing CTA at publish",                 "—", "0",   "0",    "Editorial checklist review")
r += 1

# ── 6. Review cadence ────────────────────────────────────────────────────────
r = kpi_section(ws3, r, "6. REVIEW CADENCE")
r = kpi_col_hdr(ws3, r)
r = kpi_row(ws3, r, "Weekly pipeline check",    "—", "Every Monday", "—", "Review Status column in editorial calendar. Flag any at-risk deadlines.")
r = kpi_row(ws3, r, "Mid-quarter review",       "—", "15 May 2026",  "—", "Review GSC/GA4 for first-batch articles (Apr). Adjust if needed.")
r = kpi_row(ws3, r, "End-of-quarter debrief",   "—", "15 Jul 2026",  "—", "Full KPI review. Output: Q3 calendar brief + what to replicate/drop.")

ws3.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
out = "/home/user/my-web-app/GuestReady_Q2_2026_Content_Strategy.xlsx"
wb.save(out)
print(f"Saved: {out}")
